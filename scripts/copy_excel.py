import openpyxl

input_filename = r'C:\Users\Admin\Downloads\casa employee details.xlsx'
output_filename = r"C:\Users\Admin\Downloads\Book1.xlsx"

# Load input workbook
wb_in = openpyxl.load_workbook(input_filename)
ws_in = wb_in.active

# Create output workbook
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Extracted IDs"

# Process rows starting from row 2 (skip header)
for row_num in range(2, ws_in.max_row + 1):
    cell_value = ws_in.cell(row=row_num, column=1).value  # Column A
    
    if not cell_value:
        continue  # Skip empty cells
    
    # Extract values
    first_index = cell_value.find('|')
    second_index = cell_value.find('|', first_index + 1)  # Find the second '|'
    third_index = cell_value.find('|', second_index + 1)  # Find the third '|'
    fourth_index = cell_value.find('|', third_index + 1)  # Find the fourth '|'
    
    # Extract value from 8th character to the first '|'
    value_1 = cell_value[8:first_index].strip() if first_index != -1 else cell_value[8:].strip()
    
    # Extract value from 26th character to the second '|'
    value_2 = cell_value[26:second_index].strip() if second_index != -1 else cell_value[26:].strip()

    # Extract value from second '|' + 13 character to the third '|'
    value_3 = cell_value[second_index + 13:third_index].strip() if third_index != -1 else cell_value[second_index + 13:].strip()

    # Extract value from third '|' + 14 character to the fourth '|'
    value_4 = cell_value[third_index + 14:fourth_index].strip() if fourth_index != -1 else cell_value[third_index + 14:].strip()
    
    # Write extracted value_1 to column 1
    ws_out.cell(row=row_num-1, column=1, value=value_1)  # Offset by 1 row
    
    # Write extracted value_2 to column 2
    ws_out.cell(row=row_num-1, column=2, value=value_2)  # Offset by 1 row

    # Write extracted value_3 to column 3
    ws_out.cell(row=row_num-1, column=3, value=value_3)  # Offset by 1 row

    # Write extracted value_4 to column 4
    ws_out.cell(row=row_num-1, column=4, value=value_4)  # Offset by 1 row

# Save output file
wb_out.save(output_filename)